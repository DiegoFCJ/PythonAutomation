import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import easygui

def get_product_title(isbn):
    options = webdriver.ChromeOptions()
    options.add_argument('--ignore-certificate-errors')
    driver = webdriver.Chrome(options=options)

    try:
        search_url = f"http://fabbricadeisegni.it?s={isbn}"
        driver.get(search_url)

        # Attendere che l'elemento con la classe 'aws-search-btn' sia presente
        search_button = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, "aws-search-btn"))
        )

        # Eseguire il clic utilizzando JavaScript
        driver.execute_script("arguments[0].click();", search_button)

        # Attendere che l'elemento con la classe 'product_title' sia presente nella pagina successiva
        product_title_element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, "product_title"))
        )
        # Estrai il testo dell'elemento 'product_title'
        product_title = product_title_element.text.strip()
        return product_title

    except Exception as e:
        print(f"Errore: {e}")
        return None

    finally:
        driver.quit()

# Ottenere l'ISBN dall'utente utilizzando easygui
isbn = easygui.enterbox("Inserisci l'ISBN del prodotto:")

# Ottieni il titolo del prodotto
product_title = get_product_title(isbn)

if product_title:

    # Apri il file Excel
    workbook = openpyxl.load_workbook('Leaf node Corretti.xlsx')
    sheet = workbook.active

    # Cerca l'ISBN all'interno del file Excel e scrivi il titolo del prodotto nella colonna adatta
    for row in sheet.iter_rows(min_row=2, max_col=2, max_row=sheet.max_row):
        if row[1].value is not None and str(row[1].value) == isbn:
            # Scrivi il titolo solo se l'ISBN corrisponde
            sheet.cell(row=row[1].row, column=3).value = product_title
            break

    # Salva le modifiche
    workbook.save('Leaf node Corretti.xlsx')
    print(f"Titolo scritto nel file Excel: {product_title}")
else:
    print("Impossibile trovare il titolo per l'ISBN fornito.")


