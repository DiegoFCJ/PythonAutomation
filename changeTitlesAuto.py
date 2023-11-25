import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

def get_product_title(isbn):
    options = webdriver.ChromeOptions()
    options.add_argument('--ignore-certificate-errors')
    driver = webdriver.Chrome(options=options)

    try:
        search_url = f"http://fabbricadeisegni.it?s={isbn}"
        driver.get(search_url)

        # Attendere che l'elemento con la classe 'aws-search-btn' sia presente
        search_button = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, "aws-search-btn"))
        )

        # Eseguire il clic utilizzando JavaScript
        driver.execute_script("arguments[0].click();", search_button)

        # Attendere che l'elemento con la classe 'product_title' sia presente nella pagina successiva
        product_title_element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, "product_title"))
        )
        # Estrai il testo dell'elemento 'product_title'
        product_title = product_title_element.text.strip()
        return product_title

    except Exception as e:
        print(f"Errore: {e}")
        return None

    finally:
        driver.quit()

# Apri il file Excel
workbook = openpyxl.load_workbook('Leaf node Corretti.xlsx')
sheet = workbook.active

row_number = 2  # Parti dalla seconda riga
isbn = str(sheet.cell(row=row_number, column=2).value)  # Prendi l'ISBN dalla seconda riga

while isbn is not None:
    product_title = sheet.cell(row=row_number, column=3).value
    if product_title is None:
        # Ottieni il titolo del prodotto
        product_title = get_product_title(isbn)

        if product_title:
            # Cerca l'ISBN all'interno del file Excel e scrivi il titolo del prodotto nella colonna adatta
            for row in sheet.iter_rows(min_row=2, max_col=2, max_row=sheet.max_row):
                if row[1].value is not None and str(row[1].value) == isbn:
                    print(str(row[1].value), " == ", isbn)
                    # Scrivi il titolo solo se l'ISBN corrisponde
                    sheet.cell(row=row_number, column=3).value = product_title
                    print(f"Titolo scritto nel file Excel per ISBN {isbn}: {product_title}")
                    workbook.save('Leaf node Corretti.xlsx')  # Salva immediatamente dopo l'inserimento
                    break

    row_number += 1
    isbn = str(sheet.cell(row=row_number, column=2).value)  # Passa all'ISBN successivo